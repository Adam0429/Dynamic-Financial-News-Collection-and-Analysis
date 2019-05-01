from stock_relation import _list
import xlrd,xlwt
import datetime
import pandas as pd
import pandas_datareader.data as web
from tqdm import tqdm
import IPython
import glob
import openpyxl
# import spacy



stock_prices = {}

for relations in tqdm(_list):
	try:
		price = web.DataReader(relations[0], 'yahoo')
		stock_prices[relations[0]] = price
	except:
		pass


def recent_price(stockname,start_time):
	# today = datetime.date.today()
	date = start_time.split('-')
	t1 = pd.Timestamp(datetime.datetime(int(date[0]),int(date[1]),int(date[2]),0,0,0))  # 创建一个datetime.datetime
	# High            155.789993
	# Low             154.600006
	# Open            155.570007
	# Close           155.720001
	# Volume       521800.000000
	# Adj Close       153.603760
	try:
		prices = stock_prices[stockname]
		today_index = list(prices.index).index(t1)
		tomorrow_index = prices.index[today_index+1]
		recent_indexs = [prices.index[today_index+i] for i in range(-1,2)]
		# 前一天 当天 和 后五天的 
		data = []
		for index in recent_indexs:
			data.append(prices.loc[index]['Close'])
	except KeyboardInterrupt:
		exit()
	except:
		return
	return data

df = pd.read_csv(u'news.csv',encoding='ISO-8859-1') 

titles = df['title']
contents = df['content']
dates = df['date']
# df.info()
# df.head(1)
# df[df.isnull().values==True].drop_duplicates() # 2462
# df[df.isnull().values==True].drop_duplicates()['content'] # 2462
# df['date']# 2006-10-20 . 2018-04-07 
# df['date'].describe()

_contents = set()


outwb = openpyxl.Workbook()  # 打开一个将写的文件
outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet


count = 1

for idx in tqdm(range(0,len(titles))):
	if contents[idx] in _contents or type(contents[idx]) != type('str') or type(titles[idx]) != type('str'):
			continue
	else:
		_contents.add(contents[idx])
	# 去重
	# sents = sent_tokenize(contents[idx])
	# for sent in sents:			
	for relations in _list:
		for item in relations[1:len(relations)-1]: # 不按“领域”标注公司
			if item in contents[idx]:
				recent_prices = recent_price(relations[0],dates[idx])
				if recent_prices != None:
					outws.cell(count, 1).value = titles[idx]
					outws.cell(count, 2).value = contents[idx]
					outws.cell(count, 3).value = relations[1]
					outws.cell(count, 4).value = str(recent_prices)
					outws.cell(count, 5).value = dates[idx]
					count += 1
				else:
					outws.cell(count, 1).value = titles[idx]
					outws.cell(count, 2).value = contents[idx]
					outws.cell(count, 3).value = relations[1]
					outws.cell(count, 4).value = 'None'
					outws.cell(count, 5).value = dates[idx]
					count += 1
				break


saveExcel = "labeled_data.xls"
outwb.save(saveExcel)




